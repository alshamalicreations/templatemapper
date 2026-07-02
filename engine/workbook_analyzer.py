"""
Analyzes Excel workbooks.
"""

from openpyxl.workbook.workbook import Workbook


class WorkbookAnalyzer:

    def analyze(self, workbook: Workbook):

        report = {}

        for sheet in workbook.worksheets:

            headers = []

            if sheet.max_row > 0:

                for cell in sheet[1]:
                    headers.append(cell.value)

            report[sheet.title] = {
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "headers": headers,
            }

        return report