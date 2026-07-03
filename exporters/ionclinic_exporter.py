"""
ionclinic_exporter.py

Exports clinic data into the IonClinic Simplified Import Template.
"""

from pathlib import Path

from openpyxl import load_workbook

from writers.patient_writer import PatientWriter
from writers.appointment_writer import AppointmentWriter
from writers.payment_writer import PaymentWriter


class IonClinicExporter:

    def __init__(
        self,
        template_path: str,
        clinic_id: str,
        tracker=None,
    ):

        self.template_path = Path(template_path)

        self.workbook = load_workbook(
            self.template_path
        )

        # Reserved for future exporters (Backup, etc.)
        # Not used by the Simplified Import Template.
        self.clinic_id = clinic_id

        self.tracker = tracker

        self._prepare_template()

    def _prepare_template(self):
        """
        Clears all existing data rows while preserving
        the header row of every supported worksheet.
        """

        sheets = [
            "patients",
            "appointments",
            "payments",
            "inventory_items",
            "material_usage",
        ]

        for sheet_name in sheets:

            if sheet_name not in self.workbook.sheetnames:
                continue

            sheet = self.workbook[sheet_name]

            if sheet.max_row > 1:

                sheet.delete_rows(
                    idx=2,
                    amount=sheet.max_row - 1,
                )

    def export_patients(self, patients):

        PatientWriter(
            self.workbook,
            self.tracker,
        ).write(patients)

    def export_appointments(self, patients):

        AppointmentWriter(
            self.workbook,
            self.tracker,
        ).write(patients)

    def export_payments(self, patients):

        PaymentWriter(
            self.workbook,
            self.tracker,
        ).write(patients)

    def export(self, patients):

        print("Writing patients...")
        self.export_patients(patients)
        print("✓ Patients exported")

        print()

        print("Writing appointments...")
        self.export_appointments(patients)
        print("✓ Appointments exported")

        print()

        print("Writing payments...")
        self.export_payments(patients)
        print("✓ Payments exported")

    def save(self, output_path: str):

        self.workbook.save(output_path)