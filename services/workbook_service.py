"""
workbook_service.py

Handles loading and saving Excel workbooks.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class WorkbookService:

    def load(self, file_path: str | Path) -> Workbook:

        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")

        if path.suffix.lower() != ".xlsx":
            raise ValueError("Only .xlsx files are supported.")

        return load_workbook(path)

    def save(self, workbook: Workbook, output_path: str | Path):

        workbook.save(output_path)

    def get_sheet_names(self, workbook: Workbook):

        return workbook.sheetnames

    def get_sheet(self, workbook: Workbook, sheet_name: str):

        return workbook[sheet_name]